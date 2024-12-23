from day import Day
from group_combination import GroupCombination
from week import Week
from config import get_cfg_option
import openpyxl

class DAL: #Data Abstraction Layer
    def __init__(self, filename):
        self.filename = filename
        self.groups = self.fetch_groups()
        self.schedule = self.fetch_schedule()

    def load_sheet(self, sheet_name):
        workbook = openpyxl.load_workbook(self.filename,read_only=True, data_only=True)
        return workbook[sheet_name]

    def fetch_schedule(self):
        worksheet = self.load_sheet(get_cfg_option('schedule_sheet'))
        config = get_cfg_option("schedule_sheet_info")
        rows = config["rows_per_week_including_weeks_name_rows"]
        lesson_rows = config["lessons_per_day"]
        columns = config["columns_per_work_day_except_monday"]
        monday_columns = config["columns_for_monday"]
        days = config["scheduled_days"]
        total_weeks = config["total_weeks_this_term"]
        #first column is time slots
        #firts row is a number of the week
        #no sunday and saturday

        all_weeks = []

        row_index = 0
        for row in worksheet.iter_rows(min_col = 2, max_row= total_weeks * rows, values_only=True):
            row_index += 1

            if row_index % rows == rows - lesson_rows:
                new_week = Week([Day([]) for i in range(days)])

            if (row_index - 1) % rows in range(rows - lesson_rows, rows):
                day_chunks = []
                for i in range(1,days + 1):
                    if i == 1 :
                        chunk = row[: monday_columns]
                    else:
                        chunk = row[(i - 2) * columns + monday_columns: (i - 1) * columns + monday_columns]

                    final_chunk = []
                    for cell in chunk:
                        if not cell:
                            continue

                        if not cell.strip().replace("\n", ""):
                            continue
                        group = cell.replace("\n", "").strip()
                        final_chunk.append(group)
                    day_chunks.append(final_chunk)

                # even if chunk is empty list mustn't delete because messes up days order
                day_chunks = [GroupCombination(chunk) for chunk in day_chunks]
                for i in range(days):
                    new_week.days[i].classes.append(day_chunks[i])

            if row_index % rows == 0:
                all_weeks.append(new_week)

        return all_weeks


    # returns {student_email -> their groups}
    def fetch_groups(self):
        worksheet = self.load_sheet(get_cfg_option('groups_sheet_name'))
        config = get_cfg_option("groups_sheet_column_info")
        name_column_name = config['name_column_name']
        surname_column_name = config['surname_column_name']
        columns_for_student_info = config['columns_for_student_info']

        row = list(worksheet.iter_rows(max_row=1, values_only=True))[0]
        header = [column_name for column_name in row if column_name]

        name_ind = header.index(name_column_name)
        surname_ind = header.index(surname_column_name)

        info = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            name = row[name_ind]
            surname = row[surname_ind]

            if not name and not surname:
                continue
            elif not name and surname:
                student = surname
            elif not surname and name:
                student = name
            else:
                student = surname + "_" + name

            if student not in info:  # one student can have two majors so email might happen twice
                info[student] = GroupCombination()

            groups = []
            for cell in row[columns_for_student_info::]:
                if not cell:
                    continue

                if not cell.strip().replace("\n",""):
                    continue

                group = cell.strip().replace("\n","")
                groups.append(group)

            info[student] += GroupCombination(groups)

        return info

    def write_holes_to_json(self, holes):
        pass

    def get_groups_by_student(self, student) -> GroupCombination:
        return self.groups[student]

    def get_week_by_number(self, n) -> Week:
        return self.schedule[n-1]

    def get_all_students(self):
        return self.groups.keys()

    def get_all_weeks(self):
        return self.schedule

