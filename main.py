from argparse import ArgumentParser
import logic
from dal import DAL
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule

def init_arg_parser():
    parser = ArgumentParser(prog='holefinder', description='CLI program to find the holes in the schedule') 

    parser.add_argument('filename', help='.xlsx file with the schedule information')

    return parser

def get_args():
    parser = init_arg_parser()
    return parser.parse_args()

def partition_list(lst, max_length):
    return [lst[i:i + max_length] for i in range(0, len(lst), max_length)]

def put_line_to_sheet(line, sheet, row, column):
    for i in range(len(line)):
        sheet.cell(row=row, column=column + i, value = line[i])

def save_result_to_xlsx(result):
    workbook = Workbook()
    workbook.create_sheet('Result')
    sheet = workbook['Result']

    HEADER_ROWS = 3

    put_line_to_sheet(['Total holes:', result['stats']['total_holes'], '', 'Total overlaps:', result['stats']['total_overlaps']], sheet, 1, 1)
    put_line_to_sheet(['Holes/student:', result['stats']['holes_per_student'], '', 'Overlaps/student:', result['stats']['overlaps_per_student']], sheet, 2, 1)

    key_order = sorted(result['data'].keys(), key = lambda x: result['data'][x]['total_holes'] + result['data'][x]['total_overlaps'], reverse=True)

    schedule_start_row = 0 #We need to figure this out
    column = 1
    for groups_hash in key_order: #Place info about each group combination, find where to place the actual schedule
        info = result['data'][groups_hash]
        row = HEADER_ROWS + 2

        put_line_to_sheet(['Holes:', info['total_holes']], sheet, row, column)
        row += 2

        put_line_to_sheet(['Overlaps:', info['total_overlaps']], sheet, row, column)
        row += 2

        put_line_to_sheet(['Students:', info['n_students']], sheet, row, column)
        row += 2

        sheet.cell(row=row, column=column, value='Groups:')
        row += 1

        groups = partition_list(info['groups'].groups, 5)
        for i in range(len(groups)):
            put_line_to_sheet(groups[i], sheet, row, column)
            row += 1

        row += 2
        column += 6
       
        if row > schedule_start_row: #If we are out of bounds, extend the bounds!
            schedule_start_row = row

    column = 1
    for groups_hash in key_order: #We can now safely place the schedule
        info = result['data'][groups_hash]
        row = schedule_start_row
 
        put_line_to_sheet(['Пн', 'Вт', 'Ср', 'Чт', 'Пт'], sheet, row, column)
        for week in info['schedule']:
            for i in range(8): #8 classes per day
                row += 1                    
                put_line_to_sheet([day[i] for day in week.values()], sheet, row, column)
            row += 2

        column += 6
 
    #We can now apply formatting
    start_row = schedule_start_row
    end_row = sheet.max_row
    start_col = 1
    end_col = sheet.max_column
    range_address = f"{sheet.cell(row=start_row, column=start_col).coordinate}:{sheet.cell(row=end_row, column=end_col).coordinate}"


    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid") 
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid") 
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") 

    # Conditional formatting rules
    # Rule 1: Red for 'HOLE'
    rule_red = CellIsRule(operator="equal", formula=['"HOLE"'], fill=red_fill)
    sheet.conditional_formatting.add(range_address, rule_red)

    # Rule 2: Green for 'FREE'
    rule_green = CellIsRule(operator="equal", formula=['"FREE"'], fill=green_fill)
    sheet.conditional_formatting.add(range_address, rule_green)

    # Rule 3: Orange for cells containing '|'
    rule_orange = FormulaRule(formula=[f'ISNUMBER(SEARCH("|", {sheet.cell(row=start_row, column=start_col).coordinate}))'], fill=orange_fill)
    sheet.conditional_formatting.add(range_address, rule_orange)

    rule_yellow_default = FormulaRule(formula=[f'NOT(ISBLANK({sheet.cell(row=start_row, column=start_col).coordinate}))'], fill=yellow_fill)
    sheet.conditional_formatting.add(range_address, rule_yellow_default)

    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    workbook.save('result.xlsx')

def main():
    args = get_args()
    filename = args.filename

    dal = DAL(filename)
    
    term = dal.get_all_weeks()

    result = {}
    result['data'] = {}
    students = dal.get_all_students()
    for student in students:
        groups = dal.get_groups_by_student(student)
        groups_hash = groups.get_hash()

        if groups_hash in result:
            info['n_students'] += 1 
            continue

        result['data'][groups_hash] = {}
        info = result['data'][groups_hash]


        info['groups'] = groups
        info['schedule'] = []
        info['n_students'] = 1
        info['total_holes'] = 0
        info['total_overlaps'] = 0

        for week in term:
            schedule = week.get_schedule_of(groups) 
            info['schedule'].append(schedule)
            for day in schedule.values():
                info['total_holes'] += day.count('HOLE')
                info['total_overlaps'] += len(list(filter(lambda x: '|' in x, day)))

    total_students = 0
    total_holes = 0
    total_overlaps = 0

    for groups_hash in result['data']:
        info = result['data'][groups_hash]
        total_students += info['n_students']
        total_holes += info['total_holes']
        total_overlaps += info['total_overlaps']

    result['stats'] = {}
    result['stats']['total_holes'] = total_holes
    result['stats']['total_overlaps'] = total_overlaps

    result['stats']['holes_per_student'] = total_holes/total_students    
    result['stats']['overlaps_per_student'] = total_overlaps/total_students    

    save_result_to_xlsx(result)

if __name__ == '__main__':
    main()
